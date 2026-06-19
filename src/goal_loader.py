from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO

import openpyxl
import pandas as pd

from .data_cleaning import normalize_text, normalize_udr_key
from .term_mapping import ParsedTermMetric, parse_term_metric_header


@dataclass
class GoalParseResult:
    goals: pd.DataFrame
    starts: pd.DataFrame
    current_state: dict[str, object]
    parsed_term_columns: list[str]
    parse_issues: list[str]


def _cell_text(value: object) -> str:
    return normalize_text(value)


def _as_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        return None
    return float(numeric)


def _load_workbook(source: str | Path | BinaryIO):
    return openpyxl.load_workbook(source, data_only=True, read_only=True)


def _extract_current_state(wb) -> tuple[dict[str, object], list[str]]:
    current_state: dict[str, object] = {}
    parsed_columns: list[str] = []
    for ws in wb.worksheets:
        for row_idx in range(1, ws.max_row + 1):
            row_terms = _header_term_columns(ws, row_idx)
            for parsed in row_terms.values():
                if parsed.raw_header not in parsed_columns:
                    parsed_columns.append(parsed.raw_header)

            labels = [_cell_text(ws.cell(row_idx, col_idx).value).lower() for col_idx in range(1, min(ws.max_column, 8) + 1)]
            if "budget" in labels:
                budget_idx = labels.index("budget") + 1
                value = _as_number(ws.cell(row_idx, budget_idx + 1).value)
                if value is not None and "annual_goal" not in current_state:
                    current_state["annual_goal"] = value
                    current_state["annual_goal_source"] = f"{ws.title}!{ws.cell(row_idx, budget_idx + 1).coordinate}"
            if "enrolled" in labels:
                enrolled_idx = labels.index("enrolled") + 1
                value = _as_number(ws.cell(row_idx, enrolled_idx + 1).value)
                if value is not None and "current_enrolled" not in current_state:
                    current_state["current_enrolled"] = value
                    current_state["current_enrolled_source"] = f"{ws.title}!{ws.cell(row_idx, enrolled_idx + 1).coordinate}"
            if "starts" in labels:
                starts_idx = labels.index("starts") + 1
                value = _as_number(ws.cell(row_idx, starts_idx + 1).value)
                if value is not None and "current_starts" not in current_state:
                    current_state["current_starts"] = value
                    current_state["current_starts_source"] = f"{ws.title}!{ws.cell(row_idx, starts_idx + 1).coordinate}"
            if "start%" in labels or "start %" in labels:
                start_idx = labels.index("start%") + 1 if "start%" in labels else labels.index("start %") + 1
                value = _as_number(ws.cell(row_idx, start_idx + 1).value)
                if value is not None and "current_start_pct" not in current_state:
                    current_state["current_start_pct"] = value
                    current_state["current_start_pct_source"] = f"{ws.title}!{ws.cell(row_idx, start_idx + 1).coordinate}"
    return current_state, parsed_columns


def _header_term_columns(ws, row_idx: int) -> dict[int, ParsedTermMetric]:
    columns: dict[int, ParsedTermMetric] = {}
    for col_idx in range(1, ws.max_column + 1):
        parsed = parse_term_metric_header(ws.cell(row_idx, col_idx).value)
        if parsed:
            columns[col_idx] = parsed
    return columns


def _section_name(ws, header_row: int) -> str:
    for row_idx in range(header_row - 1, max(0, header_row - 8), -1):
        values = [_cell_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        values = [value for value in values if value]
        for value in values:
            if value.lower() not in {"current state", "budget"}:
                return value
    return "Budget"


def _classify_section(section: str) -> str:
    lowered = section.lower()
    if "udr" in lowered:
        return "UDR"
    if "lead" in lowered or "spm" in lowered or "ml" in lowered:
        return "Source"
    return "Budget"


def _is_summary_name(value: str) -> bool:
    lowered = value.lower()
    return lowered in {"total", "starts", "start%", "start %", "enrolled", "leads"} or lowered.startswith("total ")


def _row_values_by_term(ws, row_idx: int, term_columns: dict[int, ParsedTermMetric]) -> dict[str, dict[str, object]]:
    by_term: dict[str, dict[str, object]] = {}
    for col_idx, parsed in term_columns.items():
        term = by_term.setdefault(
            parsed.term_code,
            {
                "term_code": parsed.term_code,
                "term_label": parsed.term_label,
                "sort_order": parsed.sort_order,
                "actual": None,
                "goal": None,
                "raw_actual_column": None,
                "raw_goal_column": None,
            },
        )
        value = _as_number(ws.cell(row_idx, col_idx).value)
        if parsed.metric == "actual":
            term["actual"] = value
            term["raw_actual_column"] = parsed.raw_header
        else:
            term["goal"] = value
            term["raw_goal_column"] = parsed.raw_header
    return by_term


def load_goal_workbook(source: str | Path | BinaryIO | None) -> GoalParseResult:
    if not source:
        return GoalParseResult(pd.DataFrame(), pd.DataFrame(), {}, [], ["No budget workbook provided."])

    try:
        wb = _load_workbook(source)
    except Exception as exc:
        return GoalParseResult(pd.DataFrame(), pd.DataFrame(), {}, [], [f"Budget workbook could not be loaded: {exc}"])

    records: list[dict[str, object]] = []
    start_records: list[dict[str, object]] = []
    parse_issues: list[str] = []
    current_state, parsed_term_columns = _extract_current_state(wb)

    for ws in wb.worksheets:
        header_rows = []
        for row_idx in range(1, ws.max_row + 1):
            term_columns = _header_term_columns(ws, row_idx)
            if len(term_columns) >= 2:
                header_rows.append((row_idx, term_columns))

        for header_index, (header_row, term_columns) in enumerate(header_rows):
            first_term_col = min(term_columns)
            name_col = max(1, first_term_col - 2)
            budget_col = max(1, first_term_col - 1)
            section = _section_name(ws, header_row)
            entity_type = _classify_section(section)
            next_header_row = header_rows[header_index + 1][0] if header_index + 1 < len(header_rows) else ws.max_row + 1

            for row_idx in range(header_row + 1, next_header_row):
                entity_name = _cell_text(ws.cell(row_idx, name_col).value)
                if not entity_name:
                    continue
                if _is_summary_name(entity_name):
                    if entity_name.lower().startswith("starts") or entity_name.lower().startswith("start"):
                        for term_record in _row_values_by_term(ws, row_idx, term_columns).values():
                            start_records.append(
                                {
                                    "sheet": ws.title,
                                    "section": section,
                                    "metric": entity_name,
                                    **term_record,
                                }
                            )
                    continue

                term_values = _row_values_by_term(ws, row_idx, term_columns)
                for term_record in term_values.values():
                    actual = term_record["actual"]
                    goal = term_record["goal"]
                    if actual is None and goal is None:
                        continue
                    variance = None if actual is None or goal is None else actual - goal
                    pct_to_goal = None if not goal else (actual or 0) / goal
                    records.append(
                        {
                            "sheet": ws.title,
                            "section": section,
                            "entity_type": entity_type,
                            "entity": entity_name,
                            "entity_key": normalize_udr_key(entity_name),
                            "budget_total": _as_number(ws.cell(row_idx, budget_col).value),
                            "variance": variance,
                            "pct_to_goal": pct_to_goal,
                            **term_record,
                        }
                    )

    goals = pd.DataFrame(records)
    starts = pd.DataFrame(start_records)
    if goals.empty:
        parse_issues.append("No actual/goal term columns were parsed from the budget workbook.")
    else:
        missing_pairs = goals[(goals["actual"].isna()) | (goals["goal"].isna())]
        if not missing_pairs.empty:
            parse_issues.append(
                f"{len(missing_pairs):,} goal rows have an unpaired actual or goal column in the budget workbook."
            )
    expected_columns = {
        "SP-1 A",
        "SP1-G",
        "SP2-A",
        "SP2-G",
        "SU1-A",
        "SU1-G",
        "SU2 -A",
        "SU2- G",
        "FA1- A",
        "FA1 -G",
        "FA2 -A",
        "FA2 -G",
    }
    missing_expected = expected_columns - set(parsed_term_columns)
    if missing_expected:
        parse_issues.append(f"Expected budget term columns not parsed: {', '.join(sorted(missing_expected))}.")
    for key, expected in [("annual_goal", 858), ("current_enrolled", 411), ("current_starts", 318)]:
        value = current_state.get(key)
        if value is None:
            parse_issues.append(f"Budget current-state value missing: {key}.")
        elif round(float(value)) != expected:
            parse_issues.append(f"Budget current-state {key} parsed as {value}, expected {expected}.")

    return GoalParseResult(goals, starts, current_state, parsed_term_columns, parse_issues)


def selected_term_goals(goals: pd.DataFrame, term_label: str | None, entity_type: str = "UDR") -> pd.DataFrame:
    if goals.empty:
        return goals
    subset = goals[goals["entity_type"].eq(entity_type)].copy()
    if term_label and term_label != "All":
        subset = subset[subset["term_label"].eq(term_label)]
    return subset
